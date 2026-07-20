"""Unit tests for TCR extractor core logic.

Run from project root:
    python -m unittest discover -s tests
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from tcr_extractor.bom_detector import find_bom_table
from tcr_extractor.header_reader import read_header_info
from tcr_extractor.utils import extract_part_number_from_text
from tcr_extractor.workbook_processor import process_workbook, validate_row


class ExtractorTests(unittest.TestCase):
    def _workbook_with_header_and_bom(self) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tool capacity report"
        ws["A1"] = "Project: Demo Project"
        ws["A2"] = "SOP"
        ws["B2"] = "2027"
        ws["A3"] = "FOTP"
        ws["B3"] = "2026-10"
        ws["A4"] = "MOB Date: 2026-09-01"
        ws["A5"] = "Prod. Plant"
        ws["B5"] = "TM"
        ws["A6"] = "EOP"
        ws["B6"] = "2032"
        ws["A7"] = "Volume"
        ws["B7"] = 100000

        # Weak fake table. BOM detection must not select this.
        ws["A10"] = "Material"
        ws["B10"] = "Weight"

        # Real BOM table.
        header_row = 20
        headers = [
            "BOM_Pos",
            "Reference PN",
            "Part Description",
            "Complexity-Category",
            "Final.Decision Board",
            "Material",
            "Material_Key_No",
            "Special Unknown Continental Column",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col).value = header

        ws.cell(row=21, column=1).value = 1
        ws.cell(row=21, column=2).value = ""
        ws.cell(row=21, column=3).value = "Bracket 290.671-01 LH black"
        ws.cell(row=21, column=4).value = "A"
        ws.cell(row=21, column=5).value = "Buy"
        ws.cell(row=21, column=6).value = "PP"
        ws.cell(row=21, column=7).value = "MK1"
        ws.cell(row=21, column=8).value = "unmapped value"
        return wb

    def test_header_extraction_aliases(self):
        wb = self._workbook_with_header_and_bom()
        header, detected_rows = read_header_info(wb.active)
        self.assertEqual(header["Project"], "Demo Project")
        self.assertEqual(header["SOP"], "2027")
        self.assertEqual(header["MOB Presentation Date"], "2026-09-01")
        self.assertEqual(header["Production Plant"], "TM")
        self.assertIn("1", detected_rows)

    def test_bom_detection_highest_confidence(self):
        wb = self._workbook_with_header_and_bom()
        header_row, mapping, unmapped, score, reasons, missing = find_bom_table(wb.active)
        self.assertEqual(header_row, 20)
        self.assertIn("Part Name", mapping)
        self.assertIn("Final Decision Board", mapping)
        self.assertGreaterEqual(score, 65)
        self.assertTrue(any("selected_highest_confidence_row=20" in r for r in reasons))
        self.assertIn(8, unmapped)

    def test_part_number_extraction_selects_continental_format(self):
        pn, cleaned = extract_part_number_from_text("Use 123456 and Bracket 290.671-01 LH black")
        self.assertEqual(pn, "290.671-01 LH")
        self.assertEqual(cleaned, "Use 123456 and Bracket black")

    def test_validation(self):
        self.assertEqual(validate_row({"Part Number": "1", "Material": "PP", "Final Decision Board": "Buy"}), "OK")
        self.assertEqual(validate_row({"Part Number": "", "Material": "PP", "Final Decision Board": "Buy"}), "Missing Part Number")
        self.assertIn("Missing Material", validate_row({"Part Number": "1", "Material": "", "Final Decision Board": ""}))

    def test_process_workbook_debug_and_unmapped(self):
        wb = self._workbook_with_header_and_bom()
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "EU 2026 Demo.xlsx"
            wb.save(path)
            result = process_workbook(path, debug_mode=True)
        self.assertTrue(result.processed_successfully)
        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].values["Part Number"], "290.671-01 LH")
        self.assertEqual(result.rows[0].values["Part Name"], "Bracket black")
        self.assertEqual(result.rows[0].values["Validation Status"], "OK")
        self.assertEqual(len(result.debug_records), 1)
        self.assertGreaterEqual(len(result.unmapped_columns), 1)


if __name__ == "__main__":
    unittest.main()
