"""Automatic BOM table detection by confidence scoring."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    BOM_COLUMN_ALIASES,
    BOM_COMBINATION_BONUSES,
    BOM_CONFIDENCE_WEIGHTS,
    BOM_SEARCH_MAX_ROWS,
    EXPECTED_BOM_COLUMNS,
    MIN_BOM_CONFIDENCE_SCORE,
)
from .utils import build_alias_lookup, display_text, is_empty, normalize_key

BOM_LOOKUP = build_alias_lookup(BOM_COLUMN_ALIASES)


@dataclass
class BomTableCandidate:
    row: int
    mapping: dict[str, int]
    unmapped_headers: dict[int, str]
    score: int
    reasons: list[str]


def _match_column(header_value: Any) -> str | None:
    """Return canonical column name for a worksheet header cell."""
    return BOM_LOOKUP.get(normalize_key(header_value))


def _score_mapping(mapping: dict[str, int]) -> tuple[int, list[str]]:
    """Calculate confidence score for a potential BOM header row."""
    score = 0
    reasons: list[str] = []
    fields = set(mapping)

    for field in sorted(fields):
        field_score = BOM_CONFIDENCE_WEIGHTS.get(field, 1)
        score += field_score
        reasons.append(f"mapped:{field}+{field_score}")

    for required_combo, bonus in BOM_COMBINATION_BONUSES:
        if required_combo.issubset(fields):
            score += bonus
            reasons.append(f"combo:{'+'.join(sorted(required_combo))}+{bonus}")

    if not ({"Part Number", "Reference Part Number", "Part Name"} & fields):
        score -= 25
        reasons.append("penalty:missing-part-identifier-25")

    return score, reasons


def _collect_header_cells(ws: Worksheet, row: int) -> tuple[dict[str, int], dict[int, str]]:
    """Map known headers and capture unknown non-empty headers from one row."""
    mapping: dict[str, int] = {}
    unmapped: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        raw = ws.cell(row=row, column=col).value
        text = display_text(raw)
        if not text:
            continue
        canonical = _match_column(text)
        if canonical and canonical not in mapping:
            mapping[canonical] = col
        elif canonical is None:
            unmapped[col] = text
    return mapping, unmapped


def find_bom_table(ws: Worksheet) -> tuple[int | None, dict[str, int], dict[int, str], int, list[str], list[str]]:
    """Find the highest-confidence BOM header row.

    Returns:
        detected row, mapped columns, unmapped header columns, score, reasons, missing expected columns
    """
    max_row = min(ws.max_row or 1, BOM_SEARCH_MAX_ROWS)
    best: BomTableCandidate | None = None

    for row in range(1, max_row + 1):
        mapping, unmapped = _collect_header_cells(ws, row)
        if not mapping:
            continue
        score, reasons = _score_mapping(mapping)
        candidate = BomTableCandidate(row=row, mapping=mapping, unmapped_headers=unmapped, score=score, reasons=reasons)
        if best is None or candidate.score > best.score:
            best = candidate

    if best is None:
        return None, {}, {}, 0, [], EXPECTED_BOM_COLUMNS.copy()

    missing = [col for col in EXPECTED_BOM_COLUMNS if col not in best.mapping]
    selected_reasons = [f"selected_highest_confidence_row={best.row}"] + best.reasons

    if best.score >= MIN_BOM_CONFIDENCE_SCORE:
        return best.row, best.mapping, best.unmapped_headers, best.score, selected_reasons, missing
    return None, {}, {}, best.score, selected_reasons, missing


def iter_bom_rows(ws: Worksheet, header_row: int, mapping: dict[str, int]):
    """Yield source row number and BOM row dictionaries after detected header row."""
    empty_row_streak = 0

    for row in range(header_row + 1, (ws.max_row or 1) + 1):
        values: dict[str, Any] = {}
        non_empty_count = 0

        for canonical_name, col in mapping.items():
            value = ws.cell(row=row, column=col).value
            values[canonical_name] = value
            if not is_empty(value):
                non_empty_count += 1

        if non_empty_count == 0:
            empty_row_streak += 1
            if empty_row_streak >= 5:
                break
            continue

        empty_row_streak = 0
        yield row, values


def iter_unmapped_values(ws: Worksheet, header_row: int, unmapped_headers: dict[int, str]):
    """Yield values from columns that were present in the BOM but were not mapped."""
    if not unmapped_headers:
        return
    empty_row_streak = 0
    for row in range(header_row + 1, (ws.max_row or 1) + 1):
        row_values = [ws.cell(row=row, column=col).value for col in unmapped_headers]
        if all(is_empty(v) for v in row_values):
            empty_row_streak += 1
            if empty_row_streak >= 5:
                break
            continue
        empty_row_streak = 0
        for col, header in unmapped_headers.items():
            value = ws.cell(row=row, column=col).value
            if not is_empty(value):
                yield row, col, header, value
