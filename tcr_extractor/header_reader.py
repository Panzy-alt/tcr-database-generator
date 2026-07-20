"""TCR header information extraction.

All worksheet labels are configured in config.py. This module only contains the
matching/search algorithm and never hardcodes business labels.
"""

from __future__ import annotations

from typing import Any
from openpyxl.worksheet.worksheet import Worksheet

from .config import HEADER_ALIASES, HEADER_SEARCH_MAX_COLUMNS, HEADER_SEARCH_MAX_ROWS
from .utils import build_alias_lookup, display_text, is_empty, normalize_key

HEADER_LOOKUP = build_alias_lookup(HEADER_ALIASES)


def _split_inline_label_value(cell_text: str) -> tuple[str, str]:
    """Return possible label/value from cells like '<label>: <value>'."""
    if not cell_text:
        return "", ""
    for sep in [":", "="]:
        if sep in cell_text:
            left, right = cell_text.split(sep, 1)
            return left.strip(), right.strip()
    return "", ""


def _nearby_value(ws: Worksheet, row: int, col: int) -> Any:
    """Find a header value near its label without relying on fixed cell positions."""
    candidates = []
    for offset in range(1, 6):
        candidates.append((row, col + offset))
    for row_offset in range(1, 4):
        for col_offset in range(0, 4):
            candidates.append((row + row_offset, col + col_offset))

    for r, c in candidates:
        if r <= 0 or c <= 0 or r > (ws.max_row or 1) or c > (ws.max_column or 1):
            continue
        value = ws.cell(row=r, column=c).value
        if not is_empty(value):
            return value
    return ""


def read_header_info(ws: Worksheet) -> tuple[dict[str, Any], str]:
    """Read configured header fields from a worksheet.

    Returns:
        header values, comma-separated detected label rows for debug reporting
    """
    found = {key: "" for key in HEADER_ALIASES}
    detected_rows: set[int] = set()
    max_row = min(ws.max_row or 1, HEADER_SEARCH_MAX_ROWS)
    max_col = min(ws.max_column or 1, HEADER_SEARCH_MAX_COLUMNS)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = display_text(ws.cell(row=row, column=col).value)
            if not text:
                continue

            field = HEADER_LOOKUP.get(normalize_key(text))
            inline_value = ""
            if field is None:
                left, right = _split_inline_label_value(text)
                if left:
                    field = HEADER_LOOKUP.get(normalize_key(left))
                    inline_value = right

            if not field or found[field] != "":
                continue

            found[field] = inline_value if inline_value else _nearby_value(ws, row, col)
            detected_rows.add(row)

    return found, ", ".join(str(r) for r in sorted(detected_rows))
